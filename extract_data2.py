import json, datetime
import openpyxl

SRC = 'data/latest.xlsx'
wb = openpyxl.load_workbook(SRC, data_only=True)

def fmt_date(d):
    return f"{d.month}/{d.day}"

def clean_num(v, coerce_slash_decimal=False):
    if v is None or v == '':
        return None
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, str):
        s = v.strip()
        # Known corrupted-cell fix: Excel mangled "33.7" into a fraction-like string "33/7"
        if coerce_slash_decimal and '/' in s:
            parts = s.split('/')
            if len(parts) == 2 and parts[0].strip().replace('.','',1).isdigit() and parts[1].strip().isdigit():
                return float(parts[0].strip() + '.' + parts[1].strip())
        try:
            return float(s)
        except ValueError:
            return None
    return None

# ---------- ROSTER ----------
ws = wb['ROSTER']
header = [c.value for c in ws[1]]
weight_date_idx = [i for i, h in enumerate(header) if i >= 5 and isinstance(h, datetime.datetime)]
weight_dates = [fmt_date(header[i]) for i in weight_date_idx]

order = []
players = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    name = row[0]
    if not name or not str(name).strip():
        continue
    pos = row[1]
    height = row[2]
    start_weight = clean_num(row[4])
    weights = [clean_num(row[i]) for i in weight_date_idx]
    order.append(name)
    players[name] = {
        'name': name,
        'pos': pos,
        'height': height,
        'startWeight': start_weight,
        'weightDates': weight_dates,
        'weights': weights,
        'character': None,
        'metrics': {},
    }

# ---------- CHARACTER EVAL ----------
ws = wb['CHARACTER EVAL']
traits = ['workEthic', 'consistency', 'coachability', 'attitude', 'toughness']
for row in ws.iter_rows(min_row=2, values_only=True):
    name = row[0]
    if not name or name not in players:
        continue
    vals = row[1:6]
    players[name]['character'] = {t: (v if v is not None else None) for t, v in zip(traits, vals)}

# ---------- METRIC SHEETS ----------
# Updated per latest workbook: '3 6 3' sheet removed, 'VERT' renamed to 'CMJ', new 'PEAK POWER'
# sheet added under SPEED.
METRIC_SHEETS = [
    ('10YD', '10 Yard Sprint', 'SPEED', 'low', 'sec'),
    ('PEAK POWER', 'Peak Power', 'SPEED', 'high', 'W/kg'),
    ('3 STEP MPH', 'MPH by Step 3', 'SPEED', 'high', 'mph'),
    ('CMJ', 'CMJ', 'BOUNCE', 'high', 'cm'),
    ('APPROACH', 'Approach Jump', 'BOUNCE', 'high', 'in'),
    ('RSI', 'RSI', 'BOUNCE', 'high', ''),
    ('TRAP DEAD', 'TBDL 1RM (.4m/s)', 'STRENGTH', 'high', 'lb'),
    ('BENCH', 'Bench 1RM', 'STRENGTH', 'high', 'lb'),
    ('CHIN UP', 'Chin Up Max Reps', 'STRENGTH', 'high', 'reps'),
    ('BODY FAT', 'Body Fat', 'FITNESS', 'low', '%'),
    ('LEAN MASS', 'Lean Mass', 'FITNESS', 'high', 'lb'),
    ('CELTIC', 'Celtic Test', 'FITNESS', 'high', ''),
]

metric_meta = {}
team_metric_avg = {}

for sheet_name, label, category, direction, unit in METRIC_SHEETS:
    ws = wb[sheet_name]
    header = [c.value for c in ws[1]]
    date_idx = [i for i, h in enumerate(header) if isinstance(h, datetime.datetime)]
    dates = [fmt_date(header[i]) for i in date_idx]
    best_idx = None
    for i, h in enumerate(header):
        if isinstance(h, str) and h.strip().upper() == 'BEST':
            best_idx = i
            break

    coerce_bug = True  # generically guard against Excel mangling e.g. "33.7" into a string "33/7"
    # TRAP DEAD/BENCH previously had a stray bodyweight-multiplier column swept into their own
    # BEST/AVERAGE formulas (fixed in this file, but harmless to keep guarding). PEAK POWER's BEST
    # cell uses =MIN() instead of =MAX() (a copy-paste bug — higher peak power is better), so its
    # own BEST cell would show the worst test instead of the best. For all three, recompute BEST
    # ourselves from only the real dated columns instead of trusting the sheet's own formula.
    recompute_best_from_series = sheet_name in ('TRAP DEAD', 'BENCH', 'PEAK POWER')

    currents = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        if not name or name not in players:
            continue
        series = [clean_num(row[i], coerce_slash_decimal=coerce_bug) for i in date_idx]
        first_v = series[0] if series else None
        current_v = series[-1] if series else None
        if recompute_best_from_series:
            clean_series = [v for v in series if v is not None]
            best_v = (max(clean_series) if direction == 'high' else min(clean_series)) if clean_series else None
        else:
            best_v = clean_num(row[best_idx], coerce_slash_decimal=coerce_bug) if best_idx is not None else None
        if best_v == 0:
            best_v = None
        players[name]['metrics'][label] = {
            'series': series,
            'first': first_v,
            'current': current_v,
            'best': best_v,
        }
        if current_v is not None:
            currents.append(current_v)

    team_avg_current = (sum(currents) / len(currents)) if currents else None
    team_metric_avg[label] = team_avg_current

    # Row 20 holds a coach-written description of what the test measures (A20="Description:", B20=text).
    description = None
    label_cell = ws.cell(row=20, column=1).value
    if isinstance(label_cell, str) and label_cell.strip().rstrip(':').upper() == 'DESCRIPTION':
        desc_val = ws.cell(row=20, column=2).value
        if isinstance(desc_val, str) and desc_val.strip():
            description = desc_val.strip()

    metric_meta[label] = {'category': category, 'direction': direction, 'unit': unit, 'dates': dates, 'description': description}

    for n in order:
        m = players[n]['metrics'].get(label)
        if m is None:
            players[n]['metrics'][label] = {'series': [], 'first': None, 'current': None, 'best': None, 'teamAvgCurrent': team_avg_current}
        else:
            m['teamAvgCurrent'] = team_avg_current

data = {
    'order': order,
    'players': players,
    'teamMetricAvg': team_metric_avg,
    'metricMeta': metric_meta,
}

with open('roster_data2.json', 'w') as f:
    json.dump(data, f)

# ---------- PHOTOS (progress photos + headshots) ----------
# These live in the PICTURES sheet as images placed with Excel's newer "Picture in Cell" / IMAGE()
# rich-data feature (Insert > Pictures > Place in Cell), NOT as normal floating/anchored pictures.
# openpyxl's regular image API (ws._images) only sees floating pictures, so these come back as
# "#VALUE!" cell errors and are otherwise invisible to it. To read them we have to walk Excel's
# rich-value chain by hand: cell vm=N (1-based) -> xl/metadata.xml futureMetadata bk[N-1]'s
# <xlrd:rvb i=".."/> (rich value index) -> xl/richData/rdrichvalue.xml rv[index]'s first <v>
# (relationship index) -> xl/richData/richValueRel.xml rel[index] r:id -> that id's target in
# xl/richData/_rels/richValueRel.xml.rels -> the actual file under xl/media/.
import re as _re, zipfile as _zipfile, base64 as _base64, datetime as _dt, io as _io
try:
    from PIL import Image as _Image
    _HAVE_PIL = True
except ImportError:
    _HAVE_PIL = False

def _sheet_rid_for_name(xlsx_path, sheet_name):
    """Map a worksheet name to its worksheets/sheetN.xml file number."""
    z = _zipfile.ZipFile(xlsx_path)
    wb_xml = z.read('xl/workbook.xml').decode()
    sheet_to_rid = dict(_re.findall(r'<sheet name="([^"]+)"[^>]*r:id="(rId\d+)"', wb_xml))
    rid = sheet_to_rid.get(sheet_name)
    if not rid:
        return None
    rels_xml = z.read('xl/_rels/workbook.xml.rels').decode()
    rid_to_num = dict(_re.findall(r'<Relationship Id="(rId\d+)"[^>]*Target="worksheets/sheet(\d+)\.xml"', rels_xml))
    return rid_to_num.get(rid)

def _extract_richvalue_images(xlsx_path, sheet_rid):
    """Returns {(row, col_letter): raw_image_bytes} for a worksheet's Picture-in-Cell images.
    Returns {} if the workbook has no rich-value image data at all (e.g. no such photos used)."""
    z = _zipfile.ZipFile(xlsx_path)
    if 'xl/richData/rdrichvalue.xml' not in z.namelist():
        return {}
    meta_xml = z.read('xl/metadata.xml').decode()
    bk_to_rv = [int(m) for m in _re.findall(r'<xlrd:rvb i="(\d+)"/>', meta_xml)]
    rv_xml = z.read('xl/richData/rdrichvalue.xml').decode()
    rv_records = _re.findall(r'<rv s="\d+">(.*?)</rv>', rv_xml)
    rv_to_relidx = [int(_re.findall(r'<v>(.*?)</v>', rec)[0]) for rec in rv_records]
    rvrel_xml = z.read('xl/richData/richValueRel.xml').decode()
    rel_ids = _re.findall(r'<rel r:id="(rId\d+)"/>', rvrel_xml)
    rels_xml = z.read('xl/richData/_rels/richValueRel.xml.rels').decode()
    rid_to_target = dict(_re.findall(r'<Relationship Id="(rId\d+)"[^>]*Target="([^"]+)"', rels_xml))

    def vm_to_bytes(vm):
        rv_idx = bk_to_rv[vm - 1]
        rel_idx = rv_to_relidx[rv_idx]
        rid = rel_ids[rel_idx]
        target = rid_to_target[rid]
        return z.read('xl/' + target.replace('../', ''))

    sheet_xml = z.read(f'xl/worksheets/sheet{sheet_rid}.xml').decode()
    result = {}
    for m in _re.finditer(r'<c r="([A-Z]+)(\d+)"([^>]*)>', sheet_xml):
        col, row, attrs = m.groups()
        vmm = _re.search(r'vm="(\d+)"', attrs)
        if not vmm:
            continue
        try:
            result[(int(row), col)] = vm_to_bytes(int(vmm.group(1)))
        except Exception:
            continue
    return result

# The workbook's Picture-in-Cell photos come out as full-resolution phone-camera PNGs (several MB
# each). Since every player's photos get baked directly into the single dashboard HTML file, that
# would balloon the page to hundreds of MB (and blow past GitHub's 100MB per-file limit) unless we
# downscale and re-encode them first. JPEG at a modest size is more than enough for a coach to
# glance at a physique-progress photo or an avatar headshot.
def _compress_image(img_bytes, max_dim=1000, quality=78):
    if not _HAVE_PIL:
        mime = 'image/png' if img_bytes[:8] == b'\x89PNG\r\n\x1a\n' else 'image/jpeg'
        return img_bytes, mime
    try:
        im = _Image.open(_io.BytesIO(img_bytes)).convert('RGB')
        w, h = im.size
        scale = min(1.0, max_dim / max(w, h))
        if scale < 1.0:
            im = im.resize((max(1, int(w * scale)), max(1, int(h * scale))), _Image.LANCZOS)
        buf = _io.BytesIO()
        im.save(buf, format='JPEG', quality=quality, optimize=True)
        return buf.getvalue(), 'image/jpeg'
    except Exception:
        # Fall back to the original bytes if PIL cannot decode this particular image.
        mime = 'image/png' if img_bytes[:8] == b'\x89PNG\r\n\x1a\n' else 'image/jpeg'
        return img_bytes, mime

def _to_data_url(img_bytes, max_dim=1000, quality=78):
    compressed, mime = _compress_image(img_bytes, max_dim=max_dim, quality=quality)
    return f'data:{mime};base64,' + _base64.b64encode(compressed).decode('ascii')

def _parse_header_date(text, year):
    if not isinstance(text, str):
        return None
    m = _re.search(r'([A-Za-z]+)\s+(\d{1,2})', text)
    if not m:
        return None
    try:
        return _dt.datetime.strptime(f'{m.group(1)} {m.group(2)} {year}', '%B %d %Y').strftime('%Y-%m-%d')
    except ValueError:
        return None

photos_data = {}
headshots_data = {}
try:
    if 'PICTURES' in wb.sheetnames:
        pics_ws = wb['PICTURES']
        sheet_rid = _sheet_rid_for_name(SRC, 'PICTURES')
        images_by_cell = _extract_richvalue_images(SRC, sheet_rid) if sheet_rid else {}
        run_date = _dt.date.today().isoformat()
        this_year = _dt.date.today().year
        before_front_date = _parse_header_date(pics_ws.cell(row=1, column=2).value, this_year) or '2000-01-01'
        before_back_date = _parse_header_date(pics_ws.cell(row=1, column=3).value, this_year) or '2000-01-01'
        after_front_hdr = _parse_header_date(pics_ws.cell(row=1, column=4).value, this_year)
        after_back_hdr = _parse_header_date(pics_ws.cell(row=1, column=5).value, this_year)
        # If the "after" column header was never updated with its own date (i.e. it is still
        # identical to the "before" header, which is the common case), fall back to today's date
        # so the after photo reliably sorts after the before photo either way.
        after_front_date = after_front_hdr if (after_front_hdr and after_front_hdr > before_front_date) else run_date
        after_back_date = after_back_hdr if (after_back_hdr and after_back_hdr > before_back_date) else run_date

        col_map = [(2, 'front', before_front_date), (3, 'back', before_back_date),
                   (4, 'front', after_front_date), (5, 'back', after_back_date)]

        for row in range(2, pics_ws.max_row + 1):
            name = pics_ws.cell(row=row, column=1).value
            if not name or name not in players:
                continue
            entries = []
            for col_idx, angle, date_str in col_map:
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                img_bytes = images_by_cell.get((row, col_letter))
                if img_bytes:
                    entries.append({'date': date_str, 'angle': angle, 'dataUrl': _to_data_url(img_bytes)})
            if entries:
                photos_data[name] = entries
            headshot_bytes = images_by_cell.get((row, 'F'))
            if headshot_bytes:
                headshots_data[name] = _to_data_url(headshot_bytes, max_dim=400, quality=75)
except Exception as e:
    print('WARNING: photo extraction failed, keeping photos/headshots empty for this run:', e)

with open('photos_data.json', 'w') as f:
    json.dump(photos_data, f)
with open('headshots_data.json', 'w') as f:
    json.dump(headshots_data, f)

print('players:', len(order))
print('photos extracted for players:', len(photos_data))
print('headshots extracted for players:', len(headshots_data))
print('Curry, Alijah CMJ series:', players['Curry, Alijah']['metrics']['CMJ']['series'])
print('Christie, Marvin Peak Power:', players['Christie, Marvin']['metrics']['Peak Power'])
print('Cunningham, Naas weightDates:', players['Cunningham, Naas']['weightDates'])
print('Cunningham, Naas weights:', players['Cunningham, Naas']['weights'])
